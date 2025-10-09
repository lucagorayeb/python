# sheets - São as páginas de uma planilha 
# Workbook - É o arquivo que contém essas páginas

import openpyxl

# Criar Planilha
planilha_test = openpyxl.Workbook()
planilha_test.create_sheet('Frutas')
planilha_test.create_sheet('Legumes')
planilha_test.create_sheet('Sementes')
print(planilha_test.sheetnames)

cabecalho_frutas = ['Título', 'Localização', 'Preço']
sheet_frutas = planilha_test['Frutas']
sheet_frutas.append(cabecalho_frutas)


frutas = [['Uva', 'Mercado', 5], ['Melancia','Mercado' ,15], ['Bolo', 'Mercado', 40]]

for fruta in frutas:
    sheet_frutas.append(fruta)


planilha_test.save('Dados supermercado.xlsx')