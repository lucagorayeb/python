import openpyxl

planilha = openpyxl.load_workbook('pessoas.xlsx')
print(planilha.sheetnames)
sheet1 = planilha['Sheet1']
print(sheet1['B4'].value)


sheet1['B4'].value = 'Jimmy'
print(sheet1['B4'].value)

for linha in sheet1.iter_rows(min_row=2, max_row=11):
    print(linha[0].value, linha[1].value, linha[2].value, linha[3].value, linha[4].value, linha[5].value, linha[6].value,)


for linha in sheet1.iter_cols(min_row=2, min_col=3 ,max_col=3):
    for cell in linha:
        print(cell.value)
