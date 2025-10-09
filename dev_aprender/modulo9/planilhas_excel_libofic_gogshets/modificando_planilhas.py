import openpyxl

planilha = openpyxl.load_workbook('pessoas.xlsx')
print(planilha.sheetnames)
sheet1 = planilha['Sheet1']
print(sheet1['C3'].value)

# Alterando o valor de uma célula da planilha
sheet1['C3'].value = 'Hashirama'
print(sheet1['C3'].value)

# Alterando o valor de uma célula da planilha
sheet1.cell(row=3, column=3, value='Mark')
print(sheet1['C3'].value)

# Percorrer uma planilha 
""" for linha in sheet1.iter_rows(min_row=2, max_row=10, min_col=2, max_col=4):
    print(linha[0].value, linha[1].value, linha[2].value) """

# Percorrer uma única coluna
for linha in sheet1.iter_cols(min_col=2, max_col=2, min_row=2):
    for cell in linha:
        print(cell.value)

